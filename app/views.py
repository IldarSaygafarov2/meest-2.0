import pandas as pd
from django.conf import settings
from django.contrib import messages
from django.core.paginator import Paginator
from django.shortcuts import render, redirect
from django.utils import datetime_safe
from openpyxl import load_workbook, Workbook

from .forms import UserRequestForm
from .models import UserRequest


def home(request):
    if request.method == "POST":
        form = UserRequestForm(data=request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Данные о получателе приняты")
            return redirect('home')
    else:
        form = UserRequestForm()

    context = {
        "form": form
    }
    return render(request, "app/index.html", context)


def elements_view(request):
    query = request.GET
    date_from = query.get('date_from')
    date_to = query.get('date_to')
    phone_number = query.get("phone_number")
    pinfl = query.get("pinfl")
    query_fullname = query.get('fullname')
    passport_series = query.get("passport_series")
    track_number = query.get("track_number")

    if query_fullname:
        elements = UserRequest.objects.filter(fullname__iregex=query_fullname)
    elif phone_number:
        elements = UserRequest.objects.filter(phone_number__iregex=phone_number)
    elif track_number:
        elements = UserRequest.objects.filter(track_number__iregex=track_number)
    elif passport_series:
        elements = UserRequest.objects.filter(passport_number__iregex=passport_series)
    elif pinfl:
        elements = UserRequest.objects.filter(pinfl__iregex=pinfl)
    elif date_from and date_to:
        elements = UserRequest.objects.filter(created_at__gte=date_from, created_at__lte=date_to)
    else:
        elements = UserRequest.objects.all()

    qs = Paginator(elements, 30)
    page = request.GET.get('page')
    elements = qs.get_page(page)

    context = {
        "elements": elements,
        "date_from": date_from if date_from else '',
        "date_to": date_to if date_to else datetime_safe.datetime.now().date(),
        "phone_number": phone_number if phone_number else '',
        "pinfl": pinfl if pinfl else '',
        "query_fullname": query_fullname if query_fullname else '',
        "passport_series": passport_series if passport_series else '',
        "track_number": track_number if track_number else '',
    }

    return render(request, "app/elements.html", context)


def get_excel(request):
    if request.method == 'POST':
        file = request.FILES.get('file-input')
        wb = load_workbook(file)
        sheet_obj = wb.active
        values = []

        result_wb = Workbook()
        result_wb.save(settings.BASE_DIR / 'app/static/test.xlsx')

        result_file = load_workbook(settings.BASE_DIR / 'app/static/test.xlsx')
        result_file_ws = result_file.active

        # print(result_file_ws)

        for i in range(1, sheet_obj.max_row + 1):
            cell_obj = sheet_obj.cell(row=i, column=1)
            cell_obj2 = sheet_obj.cell(row=i, column=2)
            values.append((cell_obj.value, cell_obj2.value))

        for value in values:
            elements = UserRequest.objects.filter(track_number=value[0])

            if elements.exists():
                elements = elements.values_list(
                    'track_number', 'phone_number', 'passport_series',
                    'passport_number', 'pinfl', 'fullname').first()
                print('by track number', elements)
            else:
                elements = UserRequest.objects.filter(phone_number__iregex=value[1]).values_list(
                    'track_number', 'phone_number', 'passport_series',
                    'passport_number', 'pinfl', 'fullname').first()
                print('by phone number', elements)

            print('final elements', elements)
            if elements is None:
                result_file_ws.append((value[0], value[1]))
                continue
            else:
                pass

            result_file_ws.append(elements)

        result_file.save(settings.BASE_DIR / 'app/static/test.xlsx')
    else:
        elements = UserRequest.objects.all().values_list('track_number', 'fullname', 'passport_series',
                                                         'passport_number', 'pinfl', 'phone_number')
        # print(elements)

    return render(request, 'app/result.html')


def save_elements_by_datetime(request, date_from='', date_to=''):
    elements = UserRequest.objects.filter(created_at__gte=date_from, created_at__lte=date_to).values_list(
        'track_number', 'fullname', 'passport_series',
        'passport_number', 'pinfl', 'phone_number')

    df = pd.DataFrame(list(elements),
                      columns=['Трек номер', 'Ф.И.О', 'Серия паспорта', 'Номер паспорта', 'ПИНФЛ', 'Номер телефона']
                      )
    df.to_excel(settings.BASE_DIR / 'app/static/datetime.xlsx')

    return render(request, 'app/result.html')
